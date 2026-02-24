"""XLSX to Markdown converter using openpyxl."""

from pathlib import Path

from .base import BaseConverter, ConversionResult
from utils.markdown_clean import clean_markdown


class XlsxConverter(BaseConverter):
    supported_extensions = (".xlsx", ".xls")

    def convert(self, source: str | Path, **kwargs) -> ConversionResult:
        path = Path(source)

        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        md_parts = [f"# {path.name}\n"]

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            md_parts.append(f"## Sheet: {sheet_name}\n")

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                md_parts.append("*Empty sheet*\n")
                continue

            # Convert cell values to strings
            str_rows = []
            for row in rows:
                str_rows.append([str(cell) if cell is not None else "" for cell in row])

            # Header
            header = str_rows[0]
            md_parts.append("| " + " | ".join(header) + " |")
            md_parts.append("| " + " | ".join("---" for _ in header) + " |")

            # Data rows
            for row in str_rows[1:]:
                padded = row + [""] * (len(header) - len(row))
                md_parts.append("| " + " | ".join(padded[:len(header)]) + " |")

            md_parts.append("")  # blank line between sheets

        wb.close()
        content = clean_markdown("\n".join(md_parts))
        return ConversionResult(content=content, title=path.name, source=str(path))
